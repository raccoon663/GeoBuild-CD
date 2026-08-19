"""Build review_manifest.xlsx from review_manifest.csv for human review.

- Keeps every diagnostic column; moves human-review fields to the front.
- Adds a relative HYPERLINK "Open Card" -> ./cards/<candidate_id>.png.
- Dropdowns (true_change / change_type / confidence / review_status),
  frozen header row, AutoFilter, conditional formatting on review_status,
  sorted by candidate_rank ascending (unranked negative controls last).
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = Path(r"outputs/human_review/review_manifest.csv")
OUT = Path(r"outputs/human_review/review_manifest.xlsx")

FRONT = [
    "candidate_id",
    "candidate_rank",
    "area_m2",
    "size_group",
    "Open Card",
    "true_change",
    "change_type",
    "confidence",
    "notes",
    "review_status",
]

TEXT_COLS = {
    "candidate_id",
    "size_group",
    "Open Card",
    "true_change",
    "change_type",
    "confidence",
    "notes",
    "review_status",
}

VALIDATION = {
    "true_change": "yes,no,uncertain",
    "change_type": "new_building,demolition,extension,roof_or_structure_change,non_building_change,uncertain",
    "confidence": "high,medium,low",
    "review_status": "pending,reviewed,recheck",
}

STATUS_FILL = {
    "pending": PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD"),
    "reviewed": PatternFill("solid", start_color="D4EDDA", end_color="D4EDDA"),
    "recheck": PatternFill("solid", start_color="F8D7DA", end_color="F8D7DA"),
}

HEADER_FILL = PatternFill("solid", start_color="305496", end_color="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WIDTHS_FRONT = {
    "candidate_id": 14,
    "candidate_rank": 11,
    "area_m2": 11,
    "size_group": 13,
    "Open Card": 11,
    "true_change": 12,
    "change_type": 30,
    "confidence": 11,
    "notes": 42,
    "review_status": 13,
}
WIDTH_DIAG = 11


def rank_key(record):
    try:
        return (float(record["candidate_rank"]), record["candidate_id"])
    except ValueError:
        return (float("inf"), record["candidate_id"])


def main() -> None:
    with open(SRC, encoding="utf-8-sig", newline="") as fh:
        records = list(csv.DictReader(fh))
    if not records:
        raise SystemExit("empty CSV")

    records.sort(key=rank_key)
    headers = FRONT + [c for c in records[0].keys() if c not in FRONT]

    wb = Workbook()
    ws = wb.active
    ws.title = "review_manifest"
    ws.append(headers)

    oc_col = headers.index("Open Card") + 1
    for rec in records:
        row = []
        for c in headers:
            v = rec.get(c, "")
            if c not in TEXT_COLS and isinstance(v, str) and v.strip() != "":
                try:
                    v = float(v)
                except ValueError:
                    pass
            row.append(v)
        row[oc_col - 1] = f'=HYPERLINK("./cards/{rec["candidate_id"]}.png","Open Card")'
        ws.append(row)

    max_row = ws.max_row
    max_col = len(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for c in headers:
        col = get_column_letter(headers.index(c) + 1)
        ws.column_dimensions[col].width = WIDTHS_FRONT.get(c, WIDTH_DIAG)

    for col_name, values in VALIDATION.items():
        if col_name not in headers:
            continue
        col = get_column_letter(headers.index(col_name) + 1)
        dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
        dv.errorTitle = "无效输入"
        dv.error = f"请从下拉列表中选择：{values}"
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max_row}")

    status_col = get_column_letter(headers.index("review_status") + 1)
    status_range = f"{status_col}2:{status_col}{max_row}"
    for value, fill in STATUS_FILL.items():
        ws.conditional_formatting.add(
            status_range, CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill)
        )

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            if cell.column_letter == get_column_letter(oc_col):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(OUT)
    print(f"rows={max_row - 1} cols={max_col}")
    print(f"OUT  -> {OUT}")


if __name__ == "__main__":
    main()
